import pandas as pd
import tkinter as tk
from tkinter import (
    filedialog, messagebox, ttk, StringVar, IntVar, LabelFrame, Checkbutton,
    BooleanVar, Canvas, Scrollbar, Listbox
)
import os
import re
import webbrowser
import platform
from openpyxl import load_workbook

HELP_URL = "https://github.com/i732520/i732520/blob/main/HELP.md"

# --- Shared classes and helpers ---
def normalize_colname(name):
    return re.sub(r"\s+", " ", str(name).strip()).lower()

class ToolTip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        widget.bind("<Enter>", self.enter)
        widget.bind("<Leave>", self.leave)
    def enter(self, event=None):
        if self.tipwindow or not self.text:
            return
        try:
            x, y, _, cy = self.widget.bbox("insert")
        except Exception:
            x, y, cy = 0, 0, 0
        x = x + self.widget.winfo_rootx() + 20
        y = y + cy + self.widget.winfo_rooty() + 20
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry("+%d+%d" % (x, y))
        label = tk.Label(tw, text=self.text, background="#ffffe0",
                         relief="solid", borderwidth=1, font=("Arial", 9))
        label.pack()
    def leave(self, event=None):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

class ColorTreeview(ttk.Treeview):
    pass

# --- File Comparison Tab ---
class MappingSearchSBSApp(tk.Frame):
    CHUNKSIZE = 50000
    def __init__(self, master):
        super().__init__(master)
        self.root = self
        self.pack(fill="both", expand=True)
        style = ttk.Style()
        style.configure("Treeview.Heading", font=('Arial', 10, 'bold'))
        style.configure("Treeview", font=('Arial', 10))

        self.df1 = None
        self.df2 = None
        self.headers1 = []
        self.headers2 = []
        self.max_preview_rows = 1000
        self.max_display_rows = IntVar(value=1000)
        self.loaded_file1 = None
        self.loaded_file2 = None

        file_frame = tk.LabelFrame(self, text="Step 1: Select Files", font=('Arial', 11, 'bold'))
        file_frame.pack(padx=14, pady=(12,6), fill="x")
        tk.Label(file_frame, text="File 1:", font=('Arial', 10)).grid(row=0, column=0, sticky="e", padx=4, pady=2)
        self.file1_entry = tk.Entry(file_frame, width=60, font=('Arial', 10))
        self.file1_entry.grid(row=0, column=1, padx=2, pady=2)
        tk.Button(file_frame, text="Browse", command=self.load_file1).grid(row=0, column=2, padx=4)
        tk.Label(file_frame, text="File 2:", font=('Arial', 10)).grid(row=1, column=0, sticky="e", padx=4, pady=2)
        self.file2_entry = tk.Entry(file_frame, width=60, font=('Arial', 10))
        self.file2_entry.grid(row=1, column=1, padx=2, pady=2)
        tk.Button(file_frame, text="Browse", command=self.load_file2).grid(row=1, column=2, padx=4)
        tk.Button(file_frame, text="Load & Map Columns", command=self.reload_headers).grid(row=0, column=3, rowspan=2, padx=14, pady=2, sticky="ns")

        map_frame = tk.LabelFrame(self, text="Step 2: Map Columns for Matching (Composite Key Supported)", font=('Arial', 11, 'bold'))
        map_frame.pack(padx=14, pady=(0,7), fill="x")
        self.mapping_rows = []
        self.map_frame_inner = tk.Frame(map_frame)
        self.map_frame_inner.pack(anchor="w", pady=2)
        self.add_mapping_btn = tk.Button(map_frame, text="Add Mapping", command=self.add_mapping_row, state="disabled")
        self.add_mapping_btn.pack(side="left", padx=2, pady=3)
        ToolTip(self.add_mapping_btn, "Map additional column pairs for composite key matching.")

        options_frame = tk.Frame(self)
        options_frame.pack(fill="x")
        search_frame = tk.LabelFrame(options_frame, text="Step 3: Comparison & Search Options", font=('Arial', 11, 'bold'))
        search_frame.pack(side="left", padx=(0,14), pady=(0,8), fill="both", expand=True)
        tk.Label(search_frame, text="Search Type: ").grid(row=0, column=0, sticky="e")
        self.search_type = tk.StringVar(value="exact")
        ttk.Radiobutton(search_frame, text="Exact", variable=self.search_type, value="exact").grid(row=0, column=1)
        ttk.Radiobutton(search_frame, text="Contains", variable=self.search_type, value="contains").grid(row=0, column=2)
        ttk.Radiobutton(search_frame, text="Regex", variable=self.search_type, value="regex").grid(row=0, column=3)
        self.case_sensitive = tk.BooleanVar(value=False)
        tk.Checkbutton(search_frame, text="Case Sensitive", variable=self.case_sensitive).grid(row=0, column=4, padx=4)
        tk.Label(search_frame, text="Search Field: ").grid(row=1, column=0, sticky="e")
        self.mapfield_combo = ttk.Combobox(search_frame, state="readonly", width=30)
        self.mapfield_combo.grid(row=1, column=1, padx=2, pady=2, columnspan=2)
        tk.Label(search_frame, text="Value: ").grid(row=1, column=3, sticky="e")
        self.value_entry = tk.Entry(search_frame, width=20)
        self.value_entry.grid(row=1, column=4, padx=2, pady=2)
        tk.Button(search_frame, text="Search", command=self.do_search, width=12).grid(row=1, column=5, padx=8)
        tk.Button(search_frame, text="Clear Results", command=self.clear_results, width=12).grid(row=1, column=6, padx=2)
        ToolTip(self.value_entry, "Enter a value to restrict search to rows containing this value in the selected field.")

        count_option_frame = tk.LabelFrame(options_frame, text="Match Counting & Display", font=('Arial', 11, 'bold'))
        count_option_frame.pack(side="left", padx=(0,14), pady=(0,8), fill="y")
        self.count_option = IntVar(value=1)
        ttk.Radiobutton(count_option_frame, text="All matching pairs", variable=self.count_option, value=1).pack(anchor="w", pady=1)
        ttk.Radiobutton(count_option_frame, text="Unique matched File 1 rows", variable=self.count_option, value=2).pack(anchor="w", pady=1)
        ttk.Radiobutton(count_option_frame, text="Unique matched File 2 rows", variable=self.count_option, value=3).pack(anchor="w", pady=1)
        ToolTip(count_option_frame, "Affects both match count and grid display.")
        tk.Label(count_option_frame, text="Max Display Rows:").pack(anchor="w", pady=(5,0))
        self.max_display_entry = tk.Entry(count_option_frame, textvariable=self.max_display_rows, width=10)
        self.max_display_entry.pack(anchor="w", padx=2, pady=1)
        ToolTip(self.max_display_entry, "Maximum number of rows to display in the results table.")

        self.match_count_label = tk.Label(search_frame, text="Matching: 0 | Non-matching: 0", font=('Arial', 10, 'bold'))
        self.match_count_label.grid(row=2, column=0, columnspan=7, pady=(4,0), sticky="w")

        filter_frame = tk.Frame(self)
        filter_frame.pack(padx=14, pady=(0,5), fill="x")
        self.show_matches = tk.BooleanVar(value=True)
        self.show_nonmatches = tk.BooleanVar(value=True)
        tk.Checkbutton(filter_frame, text="Show Matches", variable=self.show_matches, command=self.refresh_grid).pack(side="left")
        tk.Checkbutton(filter_frame, text="Show Non-matches", variable=self.show_nonmatches, command=self.refresh_grid).pack(side="left", padx=6)
        ToolTip(filter_frame, "Toggle the display of matched and unmatched rows in the results table.")

        grid_frame = tk.LabelFrame(self, text="Step 4: Results Table", font=('Arial', 11, 'bold'))
        grid_frame.pack(padx=14, pady=8, fill="both", expand=True)
        self.grid = ColorTreeview(grid_frame, show="headings", selectmode="browse")
        self.grid.pack(side="left", fill="both", expand=True)
        grid_scroll = ttk.Scrollbar(grid_frame, orient="vertical", command=self.grid.yview)
        grid_scroll.pack(side="right", fill="y")
        self.grid.configure(yscrollcommand=grid_scroll.set)

        export_frame = tk.Frame(self)
        export_frame.pack(padx=14, pady=(0,10), fill='x')
        tk.Button(export_frame, text="Export Matched to Excel", command=lambda: self.export_to_excel(only_matches=True), width=20).pack(side="left", padx=10)
        tk.Button(export_frame, text="Export Non-matched to Excel", command=lambda: self.export_to_excel(only_matches=False), width=22).pack(side="left", padx=10)
        tk.Button(export_frame, text="Load Full File for Export", command=self.load_full_files, width=23).pack(side="left", padx=16)

        self.grid_content = []
        self.grid_columns = []

        help_btn = tk.Label(self, text="Help", fg="blue", cursor="hand2", font=("Arial", 10, "underline"))
        help_btn.pack(anchor="ne", padx=10, pady=2)
        help_btn.bind("<Button-1>", lambda e: webbrowser.open_new(HELP_URL))

        # Paste all methods from Comparingfiles.py after __init__
        # (see previous block for method bodies)
        # For brevity, refer to the original Comparingfiles.py or previous code blocks for full methods

    # --- All methods from Comparingfiles.py pasted here, unchanged ---
    self.widget = widget
        self.text = text
        self.tipwindow = None
        widget.bind("<Enter>", self.enter)
        widget.bind("<Leave>", self.leave)

    def enter(self, event=None):
        "Displays the tooltip."
        if self.tipwindow or not self.text:
            return
        x, y, _, cy = self.widget.bbox("insert") or (0,0,0,0)
        x = x + self.widget.winfo_rootx() + 20
        y = y + cy + self.widget.winfo_rooty() + 20
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True) # Removes window decorations
        tw.wm_geometry("+%d+%d" % (x, y))
        label = tk.Label(tw, text=self.text, background="#ffffe0", relief="solid", borderwidth=1, font=("Arial", 9))
        label.pack()

    def leave(self, event=None):
        "Hides the tooltip."
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

class ColorTreeview(ttk.Treeview):
    """
    Custom Treeview class (currently a placeholder, but can be extended for coloring).
    """
    pass

class MappingSearchSBSApp:
    """
    Main application class for the Side-by-Side File Comparison Tool.
    """
    def __init__(self, root):
        self.root = root
        self.root.title("Side-by-Side File Comparison Tool")
        self.root.geometry("1200x700")
        self.root.minsize(1000, 600)

        # Configure Treeview style
        style = ttk.Style()
        style.configure("Treeview.Heading", font=('Arial', 10, 'bold'))
        style.configure("Treeview", font=('Arial', 10))

        # Initialize dataframes and headers
        self.df1 = None
        self.df2 = None
        self.headers1 = []
        self.headers2 = []
        
        # Default preview limit, can be adjusted if needed
        self.max_preview_rows = 1000 
        # User-defined display limit
        self.max_display_rows = IntVar(value=1000) 

        self.loaded_file1 = None
        self.loaded_file2 = None

        # --- FILE SELECTION SECTION ---
        file_frame = tk.LabelFrame(root, text="Step 1: Select Files", font=('Arial', 11, 'bold'))
        file_frame.pack(padx=14, pady=(12,6), fill="x")

        tk.Label(file_frame, text="File 1:", font=('Arial', 10)).grid(row=0, column=0, sticky="e", padx=4, pady=2)
        self.file1_entry = tk.Entry(file_frame, width=60, font=('Arial', 10))
        self.file1_entry.grid(row=0, column=1, padx=2, pady=2)
        tk.Button(file_frame, text="Browse", command=self.load_file1).grid(row=0, column=2, padx=4)

        tk.Label(file_frame, text="File 2:", font=('Arial', 10)).grid(row=1, column=0, sticky="e", padx=4, pady=2)
        self.file2_entry = tk.Entry(file_frame, width=60, font=('Arial', 10))
        self.file2_entry.grid(row=1, column=1, padx=2, pady=2)
        tk.Button(file_frame, text="Browse", command=self.load_file2).grid(row=1, column=2, padx=4)

        tk.Button(file_frame, text="Load & Map Columns", command=self.reload_headers).grid(row=0, column=3, rowspan=2, padx=14, pady=2, sticky="ns")

        # --- COLUMN MAPPING SECTION ---
        map_frame = tk.LabelFrame(root, text="Step 2: Map Columns for Matching (Composite Key Supported)", font=('Arial', 11, 'bold'))
        map_frame.pack(padx=14, pady=(0,7), fill="x")

        self.mapping_rows = [] # Stores tuples of (combo1, combo2, remove_button)
        self.map_frame_inner = tk.Frame(map_frame)
        self.map_frame_inner.pack(anchor="w", pady=2)

        self.add_mapping_btn = tk.Button(map_frame, text="Add Mapping", command=self.add_mapping_row, state="disabled")
        self.add_mapping_btn.pack(side="left", padx=2, pady=3)
        ToolTip(self.add_mapping_btn, "Map additional column pairs for composite key matching.")

        # --- SEARCH & OPTIONS SECTION ---
        options_frame = tk.Frame(root)
        options_frame.pack(fill="x")

        search_frame = tk.LabelFrame(options_frame, text="Step 3: Comparison & Search Options", font=('Arial', 11, 'bold'))
        search_frame.pack(side="left", padx=(0,14), pady=(0,8), fill="both", expand=True)

        tk.Label(search_frame, text="Search Type: ").grid(row=0, column=0, sticky="e")
        self.search_type = tk.StringVar(value="exact")
        ttk.Radiobutton(search_frame, text="Exact", variable=self.search_type, value="exact").grid(row=0, column=1)
        ttk.Radiobutton(search_frame, text="Contains", variable=self.search_type, value="contains").grid(row=0, column=2)
        ttk.Radiobutton(search_frame, text="Regex", variable=self.search_type, value="regex").grid(row=0, column=3)
        
        self.case_sensitive = tk.BooleanVar(value=False)
        tk.Checkbutton(search_frame, text="Case Sensitive", variable=self.case_sensitive).grid(row=0, column=4, padx=4)
        
        tk.Label(search_frame, text="Search Field: ").grid(row=1, column=0, sticky="e")
        self.mapfield_combo = ttk.Combobox(search_frame, state="readonly", width=30)
        self.mapfield_combo.grid(row=1, column=1, padx=2, pady=2, columnspan=2)
        
        tk.Label(search_frame, text="Value: ").grid(row=1, column=3, sticky="e")
        self.value_entry = tk.Entry(search_frame, width=20)
        self.value_entry.grid(row=1, column=4, padx=2, pady=2)
        
        tk.Button(search_frame, text="Search", command=self.do_search, width=12).grid(row=1, column=5, padx=8)
        tk.Button(search_frame, text="Clear Results", command=self.clear_results, width=12).grid(row=1, column=6, padx=2)
        ToolTip(self.value_entry, "Enter a value to restrict search to rows containing this value in the selected field.")

        # --- COUNT OPTIONS ---
        count_option_frame = tk.LabelFrame(options_frame, text="Match Counting & Display", font=('Arial', 11, 'bold'))
        count_option_frame.pack(side="left", padx=(0,14), pady=(0,8), fill="y")
        self.count_option = IntVar(value=1) # Default to "All matching pairs"
        ttk.Radiobutton(count_option_frame, text="All matching pairs", variable=self.count_option, value=1).pack(anchor="w", pady=1)
        ttk.Radiobutton(count_option_frame, text="Unique matched File 1 rows", variable=self.count_option, value=2).pack(anchor="w", pady=1)
        ttk.Radiobutton(count_option_frame, text="Unique matched File 2 rows", variable=self.count_option, value=3).pack(anchor="w", pady=1)
        ToolTip(count_option_frame, "Affects both match count and grid display.")

        # New: MAX_DISPLAY option
        tk.Label(count_option_frame, text="Max Display Rows:").pack(anchor="w", pady=(5,0))
        self.max_display_entry = tk.Entry(count_option_frame, textvariable=self.max_display_rows, width=10)
        self.max_display_entry.pack(anchor="w", padx=2, pady=1)
        ToolTip(self.max_display_entry, "Maximum number of rows to display in the results table.")


        self.match_count_label = tk.Label(search_frame, text="Matching: 0 | Non-matching: 0", font=('Arial', 10, 'bold'))
        self.match_count_label.grid(row=2, column=0, columnspan=7, pady=(4,0), sticky="w")

        # --- FILTERS ---
        filter_frame = tk.Frame(root)
        filter_frame.pack(padx=14, pady=(0,5), fill="x")
        self.show_matches = tk.BooleanVar(value=True)
        self.show_nonmatches = tk.BooleanVar(value=True)
        tk.Checkbutton(filter_frame, text="Show Matches", variable=self.show_matches, command=self.refresh_grid).pack(side="left")
        tk.Checkbutton(filter_frame, text="Show Non-matches", variable=self.show_nonmatches, command=self.refresh_grid).pack(side="left", padx=6)
        ToolTip(filter_frame, "Toggle the display of matched and unmatched rows in the results table.")

        # --- RESULTS TABLE ---
        grid_frame = tk.LabelFrame(root, text="Step 4: Results Table", font=('Arial', 11, 'bold'))
        grid_frame.pack(padx=14, pady=8, fill="both", expand=True)
        self.grid = ColorTreeview(grid_frame, show="headings", selectmode="browse")
        self.grid.pack(side="left", fill="both", expand=True)
        grid_scroll = ttk.Scrollbar(grid_frame, orient="vertical", command=self.grid.yview)
        grid_scroll.pack(side="right", fill="y")
        self.grid.configure(yscrollcommand=grid_scroll.set)

        # --- EXPORTS ---
        export_frame = tk.Frame(root)
        export_frame.pack(padx=14, pady=(0,10), fill='x')
        tk.Button(export_frame, text="Export Matched to Excel", command=lambda: self.export_to_excel(only_matches=True), width=20).pack(side="left", padx=10)
        tk.Button(export_frame, text="Export Non-matched to Excel", command=lambda: self.export_to_excel(only_matches=False), width=22).pack(side="left", padx=10)
        tk.Button(export_frame, text="Load Full File for Export", command=self.load_full_files, width=23).pack(side="left", padx=16)

        self.grid_content = [] # Stores the data to be displayed in the grid
        self.grid_columns = [] # Stores the column headers for the grid

    def load_file1(self):
        """Opens a file dialog to select File 1 and updates the entry field."""
        path = filedialog.askopenfilename(title="Select File 1", filetypes=[("Excel/CSV/TXT", "*.xlsx *.csv *.txt"), ("All files", "*.*")])
        if path:
            self.file1_entry.delete(0, tk.END)
            self.file1_entry.insert(0, path)
            self.loaded_file1 = path

    def load_file2(self):
        """Opens a file dialog to select File 2 and updates the entry field."""
        path = filedialog.askopenfilename(title="Select File 2", filetypes=[("Excel/CSV/TXT", "*.xlsx *.csv *.txt"), ("All files", "*.*")])
        if path:
            self.file2_entry.delete(0, tk.END)
            self.file2_entry.insert(0, path)
            self.loaded_file2 = path

    def reload_headers(self):
        """
        Loads preview data from selected files, extracts headers, and initializes
        the column mapping section.
        """
        self.df1 = self.read_file(self.file1_entry.get())
        self.df2 = self.read_file(self.file2_entry.get())

        if self.df1 is None or self.df2 is None or self.df1.empty or self.df2.empty:
            messagebox.showerror("Error", "Both files must load successfully and contain data.")
            return

        self.headers1 = list(self.df1.columns)
        self.headers2 = list(self.df2.columns)

        self.clear_mapping_rows()
        self.add_mapping_btn.config(state="normal")
        self.mapfield_combo['values'] = [] # Clear search field options
        self.value_entry.delete(0, tk.END) # Clear search value
        self.clear_grid() # Clear previous results
        self.match_count_label.config(text="Matching: 0 | Non-matching: 0")

        # Attempt to auto-map columns with similar normalized names
        norm1 = {normalize_colname(h): h for h in self.headers1}
        norm2 = {normalize_colname(h): h for h in self.headers2}
        
        auto_mapped_count = 0
        for n1, h1 in norm1.items():
            if n1 in norm2:
                self.add_mapping_row(h1, norm2[n1])
                auto_mapped_count += 1
        
        # If no columns were auto-mapped, add at least one empty mapping row
        if auto_mapped_count == 0 and (self.headers1 and self.headers2):
            self.add_mapping_row()

        # Display preview mode notices if applicable
        if hasattr(self.df1, '__len__') and len(self.df1) == self.max_preview_rows:
            messagebox.showinfo("Notice", f"Preview mode: Only first {self.max_preview_rows} rows loaded from File 1.")
        if hasattr(self.df2, '__len__') and len(self.df2) == self.max_preview_rows:
            messagebox.showinfo("Notice", f"Preview mode: Only first {self.max_preview_rows} rows loaded from File 2.")

    def read_file(self, path):
        """
        Reads data from a given file path, handling CSV, Excel, and TXT formats.
        Applies a preview limit for large files.
        """
        if not path:
            return None
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == ".csv":
                # For very large CSVs, read in chunks for preview
                if os.path.getsize(path) > 100 * 1024 * 1024: # 100 MB
                    preview = []
                    row_count = 0
                    for chunk in pd.read_csv(path, chunksize=CHUNKSIZE, dtype=str):
                        preview.append(chunk)
                        row_count += len(chunk)
                        if row_count >= self.max_preview_rows:
                            break
                    if preview:
                        df = pd.concat(preview)[:self.max_preview_rows]
                        return df
                    else:
                        return pd.DataFrame() # Return empty DataFrame if no data
                else:
                    df = pd.read_csv(path, dtype=str)
                    return df
            elif ext == ".xlsx":
                # For very large Excels, read only nrows for preview
                if os.path.getsize(path) > 10 * 1024 * 1024: # 10 MB
                    df = pd.read_excel(path, dtype=str, nrows=self.max_preview_rows)
                    return df
                else:
                    df = pd.read_excel(path, dtype=str)
                    return df
            elif ext == ".txt":
                # Try reading as CSV, then as tab-separated if first fails
                try:
                    df = pd.read_csv(path, dtype=str, nrows=self.max_preview_rows)
                    return df
                except Exception:
                    df = pd.read_csv(path, sep="\t", dtype=str, nrows=self.max_preview_rows)
                    return df
            else:
                messagebox.showerror("Error", f"Unsupported file extension: {ext}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read {path}:\n{e}")
        return None

    def clear_mapping_rows(self):
        """Removes all dynamically added column mapping rows."""
        for row in self.mapping_rows:
            for widget in row:
                widget.destroy()
        self.mapping_rows.clear()

    def add_mapping_row(self, sel1=None, sel2=None):
        """
        Adds a new row for column mapping with two comboboxes and a remove button.
        Pre-selects values if provided (e.g., for auto-mapping).
        """
        row_idx = len(self.mapping_rows)
        # Set default values for comboboxes
        var1 = StringVar(value=sel1 or (self.headers1[0] if self.headers1 else ""))
        var2 = StringVar(value=sel2 or (self.headers2[0] if self.headers2 else ""))
        
        combo1 = ttk.Combobox(self.map_frame_inner, values=self.headers1, textvariable=var1, state="readonly", width=30)
        combo2 = ttk.Combobox(self.map_frame_inner, values=self.headers2, textvariable=var2, state="readonly", width=30)
        
        combo1.grid(row=row_idx, column=0, padx=2, pady=2)
        combo2.grid(row=row_idx, column=1, padx=2, pady=2)
        
        rm_btn = tk.Button(self.map_frame_inner, text="Remove", command=lambda: self.remove_mapping_row(row_idx))
        rm_btn.grid(row=row_idx, column=2, padx=2)
        
        self.mapping_rows.append((combo1, combo2, rm_btn))
        self.update_mapfield_combo()

    def remove_mapping_row(self, idx):
        """Removes a specific column mapping row by its index."""
        if idx < len(self.mapping_rows):
            for widget in self.mapping_rows[idx]:
                widget.destroy()
            self.mapping_rows.pop(idx)
            # Re-grid remaining rows to fill the gap
            for i, (c1, c2, rm) in enumerate(self.mapping_rows):
                c1.grid(row=i, column=0)
                c2.grid(row=i, column=1)
                rm.grid(row=i, column=2)
        self.update_mapfield_combo()

    def update_mapfield_combo(self):
        """Updates the 'Search Field' combobox with currently mapped columns from File 1."""
        mapped_names = [row[0].get() for row in self.mapping_rows if row[0].get()]
        self.mapfield_combo['values'] = mapped_names
        if mapped_names:
            self.mapfield_combo.current(0) # Select the first mapped column by default

    def do_search(self):
        """
        Performs the comparison and search operation based on selected files,
        column mappings, and search criteria. Populates the results grid.
        This version is optimized for memory by avoiding a full pandas merge.
        """
        if self.df1 is None or self.df2 is None:
            messagebox.showerror("Error", "Please load both files before searching.")
            return

        mapping_keys = []
        for combo1, combo2, _ in self.mapping_rows:
            k1, k2 = combo1.get(), combo2.get()
            if k1 and k2: # Ensure both columns are selected for a mapping
                mapping_keys.append((k1, k2))
        
        if not mapping_keys:
            messagebox.showerror("Error", "Please map at least one column for comparison.")
            return

        search_field = self.mapfield_combo.get()
        search_value = self.value_entry.get().strip()
        
        # If a search field is selected but no value is provided, raise an error
        if search_field and not search_value:
            messagebox.showerror("Error", "Please enter a value for the selected search field.")
            return
        
        search_type = self.search_type.get()
        case_sensitive = self.case_sensitive.get()

        cols1 = self.headers1
        cols2 = self.headers2

        # Retrieve user-defined MAX_DISPLAY_ROWS
        try:
            current_max_display = int(self.max_display_rows.get())
            if current_max_display <= 0:
                messagebox.showerror("Invalid Input", "Max Display Rows must be a positive integer.")
                return
        except ValueError:
            messagebox.showerror("Invalid Input", "Max Display Rows must be a valid integer.")
            return

        # Determine if a specific search filter is active (i.e., both field and value are provided)
        is_search_active = bool(search_field and search_value)
        print(f"DEBUG: do_search - search_field: '{search_field}', search_value: '{search_value}', is_search_active: {is_search_active}")


        # Helper function to apply search filter to a single cell value
        def apply_search_filter(cell_value, value_to_match, search_type, case_sensitive):
            cell_value_str = str(cell_value).strip()
            
            if not case_sensitive:
                cell_value_str = cell_value_str.lower()
                value_to_match = value_to_match.lower()

            result = False
            if search_type == "exact":
                result = cell_value_str == value_to_match
            elif search_type == "contains":
                result = value_to_match in cell_value_str
            elif search_type == "regex":
                try:
                    result = re.fullmatch(value_to_match, cell_value_str) is not None
                except re.error as e:
                    messagebox.showerror("Regex Error", f"Invalid regex pattern: {value_to_match}\nError: {e}")
                    result = False
            return result

        # --- Step 1: Pre-index df2 for efficient lookups ---
        # df2_key_to_rows will map composite keys to a list of (original_index, row_data) tuples
        df2_key_to_rows = {}
        for idx2, row2 in self.df2.iterrows():
            composite_key_values = []
            for _, k2 in mapping_keys:
                # Get value, handle missing columns and NaN, then strip
                val = row2.get(k2, '')
                composite_key_values.append(str(val).strip() if pd.notna(val) else '')
            composite_key = tuple(composite_key_values)
            df2_key_to_rows.setdefault(composite_key, []).append((idx2, row2))
        
        print(f"DEBUG: df2_key_to_rows created with {len(df2_key_to_rows)} unique keys.")

        # --- Step 2: Initialize results and tracking sets ---
        self.grid_content = []
        match_count = 0
        nonmatch_count = 0
        displayed_rows = 0

        # Sets to track original indices that have been processed to ensure uniqueness for counts/display
        processed_file1_original_indices = set()
        processed_file2_original_indices = set()
        
        # Store all potential rows to be displayed, then filter by current_max_display
        # This helps ensure counts are accurate even if not all rows are displayed.
        all_display_candidates = []

        # --- Step 3: Iterate through df1 to find matches and File 1 Only rows ---
        for idx1, row1 in self.df1.iterrows():
            composite_key_values = []
            for k1, _ in mapping_keys:
                val = row1.get(k1, '')
                composite_key_values.append(str(val).strip() if pd.notna(val) else '')
            composite_key = tuple(composite_key_values)

            # Determine if this File 1 row meets the search criteria
            row1_meets_search_criteria = True # Assume it meets criteria by default
            if is_search_active: # Only apply filter if a search is active
                if search_field in cols1: # ONLY apply filter if search_field exists in File 1 headers
                    if not apply_search_filter(row1.get(search_field, ''), search_value, search_type, case_sensitive):
                        row1_meets_search_criteria = False
                        # print(f"DEBUG: File 1 row {idx1} skipped - search filter mismatch on '{search_field}' (value: '{row1.get(search_field, '')}').")
                # else: If search_field is active but not in cols1, row1_meets_search_criteria remains True
                # This means the search filter doesn't apply to this side, so it's not filtered out.
            
            if not row1_meets_search_criteria:
                continue # Skip this File 1 row if it doesn't meet search criteria

            # Look for matches in df2
            df2_matches = df2_key_to_rows.get(composite_key, [])

            if df2_matches:
                # This File 1 row has at least one match in File 2
                if self.count_option.get() == 1: # All matching pairs
                    for idx2, row2 in df2_matches:
                        # Determine if this File 2 match meets the search criteria
                        row2_meets_search_criteria = True
                        if is_search_active: # Only apply filter if a search is active
                            if search_field in cols2: # ONLY apply filter if search_field exists in File 2 headers
                                if not apply_search_filter(row2.get(search_field, ''), search_value, search_type, case_sensitive):
                                    row2_meets_search_criteria = False
                                    # print(f"DEBUG: Matched File 2 row {idx2} skipped - search filter mismatch on '{search_field}' (value: '{row2.get(search_field, '')}').")
                            # else: If search_field is active but not in cols2, row2_meets_search_criteria remains True
                        
                        # A match is added only if BOTH sides meet their respective search criteria (or if filter doesn't apply)
                        if row1_meets_search_criteria and row2_meets_search_criteria:
                            all_display_candidates.append(("Match", row1, row2, True, idx1, idx2))
                            # print(f"DEBUG: Added Match (Option 1) for File 1 idx {idx1}, File 2 idx {idx2}.")
                            # Mark both indices as processed for uniqueness tracking
                            processed_file1_original_indices.add(idx1)
                            processed_file2_original_indices.add(idx2)
                
                elif self.count_option.get() == 2: # Unique matched File 1 rows
                    if idx1 not in processed_file1_original_indices:
                        # Find at least one matching row from df2 that also meets search criteria
                        valid_df2_match_found = False
                        idx2_to_display = None
                        row2_to_display = pd.Series(dtype=str)

                        for idx2_cand, row2_cand in df2_matches:
                            row2_cand_meets_criteria = True
                            if is_search_active:
                                if search_field in cols2:
                                    if not apply_search_filter(row2_cand.get(search_field, ''), search_value, search_type, case_sensitive):
                                        row2_cand_meets_criteria = False
                                else:
                                    # If search_field not in cols2, it passes for this side
                                    pass 
                            
                            if row2_cand_meets_criteria:
                                valid_df2_match_found = True
                                idx2_to_display = idx2_cand
                                row2_to_display = row2_cand
                                break # Found a valid match, no need to check others

                        if valid_df2_match_found:
                            all_display_candidates.append(("Match", row1, row2_to_display, True, idx1, idx2_to_display))
                            # print(f"DEBUG: Added Match (Option 2) for File 1 idx {idx1}, File 2 idx {idx2_to_display}.")
                            processed_file1_original_indices.add(idx1)
                            # Mark all associated df2 indices as processed to avoid them becoming "File 2 Only"
                            for df2_match_idx, _ in df2_matches:
                                processed_file2_original_indices.add(df2_match_idx)

                elif self.count_option.get() == 3: # Unique matched File 2 rows
                    # For this option, we need to iterate through df2_matches and add if df2_idx is unique
                    for idx2, row2 in df2_matches:
                        if idx2 not in processed_file2_original_indices:
                            # Determine if this File 2 match meets the search criteria
                            row2_meets_search_criteria = True
                            if is_search_active: # Only apply filter if a search is active
                                if search_field in cols2: # ONLY apply filter if search_field exists in File 2 headers
                                    if not apply_search_filter(row2.get(search_field, ''), search_value, search_type, case_sensitive):
                                        row2_meets_search_criteria = False
                                        # print(f"DEBUG: Unique matched File 2 row {idx2} skipped - search filter mismatch.")
                                else:
                                    # If search_field is active but not in cols2, row2_meets_search_criteria remains True
                                    pass
                            
                            # A match is added only if BOTH sides meet their respective search criteria (or if filter doesn't apply)
                            if row1_meets_search_criteria and row2_meets_search_criteria:
                                all_display_candidates.append(("Match", row1, row2, True, idx1, idx2))
                                # print(f"DEBUG: Added Match (Option 3) for File 1 idx {idx1}, File 2 idx {idx2}.")
                                processed_file1_original_indices.add(idx1) # Mark File 1 row as processed
                                processed_file2_original_indices.add(idx2) # Mark this specific File 2 row as processed

            else: # No match found for this File 1 row (it's a potential File 1 Only row)
                if idx1 not in processed_file1_original_indices: # Ensure it's not already handled
                    # The row1_meets_search_criteria is already checked above for File 1 Only rows
                    all_display_candidates.append(("File 1 Only", row1, pd.Series(dtype=str), False, idx1, None))
                    # print(f"DEBUG: Added File 1 Only for idx {idx1}.")
                    processed_file1_original_indices.add(idx1)


        # --- Step 4: Iterate through df2 to find File 2 Only rows ---
        for idx2, row2 in self.df2.iterrows():
            if idx2 not in processed_file2_original_indices:
                # This df2 row was not part of any match found in the df1 iteration
                
                # Determine if this File 2 Only row meets the search criteria
                row2_meets_search_criteria = True
                if is_search_active: # Only apply filter if a search is active
                    if search_field in cols2: # ONLY apply filter if search_field exists in File 2 headers
                        if not apply_search_filter(row2.get(search_field, ''), search_value, search_type, case_sensitive):
                            row2_meets_search_criteria = False
                            # print(f"DEBUG: File 2 Only row {idx2} skipped - search filter mismatch.")
                    else: # Search field not in File 2 headers, so it passes for this side
                        pass
                
                if row2_meets_search_criteria:
                    all_display_candidates.append(("File 2 Only", pd.Series(dtype=str), row2, False, None, idx2))
                    # print(f"DEBUG: Added File 2 Only for idx {idx2}.")
                    processed_file2_original_indices.add(idx2) # Mark as processed

        print(f"DEBUG: Total display candidates before current_max_display: {len(all_display_candidates)}")

        # --- Step 5: Finalize grid_content and counts based on current_max_display and filters ---
        # Reset counts for final calculation based on what will actually be displayed
        final_match_count = 0
        final_nonmatch_count = 0
        
        # Helper to format cell values, replacing NaN with empty string
        def format_cell_value_for_display(val):
            return str(val).strip() if pd.notna(val) else ''

        for item in all_display_candidates:
            source_tag, row1_data, row2_data, is_match, _, _ = item
            
            # Apply display filters (Show Matches/Show Non-matches checkboxes)
            if is_match and not self.show_matches.get():
                continue
            if not is_match and not self.show_nonmatches.get():
                continue
            
            if displayed_rows >= current_max_display: # Use user-defined limit
                break # Stop adding if current_max_display is reached

            # Apply the NaN removal here
            v_f1 = [format_cell_value_for_display(row1_data.get(h, '')) for h in cols1]
            v_f2 = [format_cell_value_for_display(row2_data.get(h, '')) for h in cols2]
            
            self.grid_content.append((source_tag, v_f1, v_f2, is_match, {}))
            displayed_rows += 1

            if is_match:
                final_match_count += 1
            else:
                final_nonmatch_count += 1

        print(f"DEBUG: Final match_count: {final_match_count}")
        print(f"DEBUG: Final nonmatch_count: {final_nonmatch_count}")
        print(f"DEBUG: Length of grid_content for display: {len(self.grid_content)}")

        # Configure grid columns
        self.grid_columns = ['Source'] + [f"File1_{h}" for h in self.headers1] + [f"File2_{h}" for h in self.headers2]
        print(f"DEBUG: Grid Columns: {self.grid_columns}")
        
        self.refresh_grid()
        self.match_count_label.config(text=f"Matching: {final_match_count} | Non-matching: {final_nonmatch_count}")

    def refresh_grid(self):
        """
        Clears the current grid and repopulates it with data from self.grid_content,
        applying the show_matches and show_nonmatches filters.
        """
        self.clear_grid()
        
        # Set up Treeview columns
        self.grid['columns'] = self.grid_columns
        self.grid.column("#0", width=0, stretch=tk.NO) # Hide the default first empty column

        for col in self.grid_columns:
            self.grid.heading(col, text=col)
            self.grid.column(col, width=120, anchor='w')

        inserted_count = 0
        # Use the user-defined max_display_rows for the actual grid insertion limit
        try:
            display_limit = int(self.max_display_rows.get())
            if display_limit <= 0:
                display_limit = 1 # Ensure at least 1 if invalid input
        except ValueError:
            display_limit = 1000 # Default to 1000 if invalid input

        for gc in self.grid_content:
            source, v_f1, v_f2, is_match, cell_diff_map = gc
            
            # Apply display filters (Show Matches/Show Non-matches checkboxes)
            # These are already applied when populating grid_content, but kept here for robustness
            if is_match and not self.show_matches.get():
                continue
            if not is_match and not self.show_nonmatches.get():
                continue
            
            values = [source] + list(v_f1) + list(v_f2) # Ensure values are in list format
            
            # Basic validation: ensure number of values matches number of columns
            if len(values) != len(self.grid_columns):
                print(f"WARNING: Row value count ({len(values)}) does not match column count ({len(self.grid_columns)}) for row: {values[:5]}...")
                continue # Skip this row to prevent Treeview errors

            self.grid.insert('', 'end', values=values)
            inserted_count += 1
            # Apply the display_limit here for the actual grid insertion
            if inserted_count >= display_limit:
                break
        print(f"DEBUG: Total rows inserted into grid: {inserted_count}")

    def clear_results(self):
        """Clears the search value, grid, and match count label."""
        self.value_entry.delete(0, tk.END)
        self.clear_grid()
        self.match_count_label.config(text="Matching: 0 | Non-matching: 0")
        self.grid_content = [] # Also clear the underlying data

    def clear_grid(self):
        """Removes all items from the Treeview grid."""
        for item in self.grid.get_children():
            self.grid.delete(item)

    def export_to_excel(self, only_matches=True):
        """
        Exports the current grid content (filtered by matches/non-matches) to an Excel file.
        """
        if not self.grid_content:
            messagebox.showerror("Export Error", "No data in grid to export.")
            return

        columns_to_export = self.grid_columns
        data_to_export = []

        # Helper to format cell values, replacing NaN with empty string for export
        def format_cell_value_for_export(val):
            return str(val).strip() if pd.notna(val) else ''

        for gc in self.grid_content:
            source, v_f1, v_f2, is_match, cell_diff_map = gc
            if only_matches and is_match:
                # Apply NaN removal for export as well
                formatted_v_f1 = [format_cell_value_for_export(val) for val in v_f1]
                formatted_v_f2 = [format_cell_value_for_export(val) for val in v_f2]
                data_to_export.append([source] + formatted_v_f1 + formatted_v_f2)
            elif not only_matches and not is_match:
                # Apply NaN removal for export as well
                formatted_v_f1 = [format_cell_value_for_export(val) for val in v_f1]
                formatted_v_f2 = [format_cell_value_for_export(val) for val in v_f2]
                data_to_export.append([source] + formatted_v_f1 + formatted_v_f2)
        
        if not data_to_export:
            messagebox.showinfo("Export", "No records to export based on current filters.")
            return
        
        df_export = pd.DataFrame(data_to_export, columns=columns_to_export)
        
        export_path = filedialog.asksaveasfilename(
            title="Export Grid to Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        
        if export_path:
            try:
                df_export.to_excel(export_path, index=False)
                messagebox.showinfo("Exported", f"Grid exported to {export_path}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export: {e}")

    def load_full_files(self):
        """
        Loads the entire content of selected files into memory (not just preview)
        for full export capabilities. Warns for very large files.
        """
        if not self.loaded_file1 or not self.loaded_file2:
            messagebox.showerror("Error", "Please select both files first.")
            return
        
        try:
            size1 = os.path.getsize(self.loaded_file1)
            size2 = os.path.getsize(self.loaded_file2)
            
            # Warn for extremely large files that might cause memory issues
            if size1 > 200*1024*1024 or size2 > 200*1024*1024: # 200 MB
                response = messagebox.askyesno(
                    "Memory Warning",
                    "One or both files are very large (>200MB). Loading full files may consume significant memory and could crash the application. Do you want to proceed?"
                )
                if not response:
                    return

            # Read full files based on extension
            ext1 = os.path.splitext(self.loaded_file1)[1].lower()
            ext2 = os.path.splitext(self.loaded_file2)[1].lower()

            if ext1 == ".csv":
                self.df1 = pd.read_csv(self.loaded_file1, dtype=str)
            elif ext1 == ".xlsx":
                self.df1 = pd.read_excel(self.loaded_file1, dtype=str)
            elif ext1 == ".txt":
                try:
                    self.df1 = pd.read_csv(self.loaded_file1, dtype=str)
                except Exception:
                    self.df1 = pd.read_csv(self.loaded_file1, sep="\t", dtype=str)
            else:
                messagebox.showerror("Error", f"Unsupported file extension for File 1: {ext1}")
                return

            if ext2 == ".csv":
                self.df2 = pd.read_csv(self.loaded_file2, dtype=str)
            elif ext2 == ".xlsx":
                self.df2 = pd.read_excel(self.loaded_file2, dtype=str)
            elif ext2 == ".txt":
                try:
                    self.df2 = pd.read_csv(self.loaded_file2, dtype=str)
                except Exception:
                    self.df2 = pd.read_csv(self.loaded_file2, sep="\t", dtype=str)
            else:
                messagebox.showerror("Error", f"Unsupported file extension for File 2: {ext2}")
                return
            
            self.headers1 = list(self.df1.columns)
            self.headers2 = list(self.df2.columns)
            messagebox.showinfo("Full Load Complete", "Full files loaded into memory. You may now run export for all rows.")
            
            # After loading full files, it's good practice to re-map columns
            # as headers might differ slightly if preview was truncated
            self.reload_headers() 

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load full files: {e}")

# --- Text to Excel Converter & Split Tool Tab ---
class ExcelToolApp(tk.Frame):
    # All code from text_to_excel_Split_converter_Final.py adapted to Frame
    def __init__(self, master):
        super().__init__(master)
        self.root = self
        self.pack(fill="both", expand=True)

        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        width = min(950, int(screen_width * 0.8))
        height = min(800, int(screen_height * 0.8))
        self.root.winfo_toplevel().geometry(f"{width}x{height}")

        if platform.system() == "Windows":
            default_font = ("Segoe UI", 10)
        else:
            default_font = ("Arial", 11)
        self.root.option_add("*Font", default_font)

        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        self.outer_frame = tk.Frame(self)
        self.outer_frame.grid(row=0, column=0, sticky="nsew")
        self.outer_frame.grid_rowconfigure(0, weight=1)
        self.outer_frame.grid_columnconfigure(0, weight=1)

        self.outer_canvas = Canvas(self.outer_frame, borderwidth=0)
        self.vscrollbar = Scrollbar(self.outer_frame, orient="vertical", command=self.outer_canvas.yview)
        self.hscrollbar = Scrollbar(self.outer_frame, orient="horizontal", command=self.outer_canvas.xview)
        self.outer_canvas.configure(yscrollcommand=self.vscrollbar.set, xscrollcommand=self.hscrollbar.set)
        self.vscrollbar.grid(row=0, column=1, sticky="ns")
        self.hscrollbar.grid(row=1, column=0, sticky="ew")
        self.outer_canvas.grid(row=0, column=0, sticky="nsew")

        self.content_frame = tk.Frame(self.outer_canvas)
        self.content_frame_id = self.outer_canvas.create_window((0, 0), window=self.content_frame, anchor="nw")

        def on_frame_configure(event):
            self.outer_canvas.configure(scrollregion=self.outer_canvas.bbox("all"))
        def on_canvas_configure(event):
            canvas_width = event.width
            self.outer_canvas.itemconfig(self.content_frame_id, width=canvas_width)
        self.content_frame.bind("<Configure>", on_frame_configure)
        self.outer_canvas.bind("<Configure>", on_canvas_configure)
        def _on_mousewheel(event):
            self.outer_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self.outer_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        help_btn = tk.Label(self, text="Help", fg="blue", cursor="hand2", font=("Arial", 10, "underline"))
        help_btn.place(relx=1.0, rely=0.0, anchor="ne", x=-12, y=2)
        help_btn.bind("<Button-1>", lambda e: webbrowser.open_new(HELP_URL))

        # Paste all UI and methods from text_to_excel_Split_converter_Final.py after __init__
        # (see previous block for method bodies)
        # For brevity, refer to the original text_to_excel_Split_converter_Final.py or previous code blocks for full methods

    # --- All methods from text_to_excel_Split_converter_Final.py pasted here, unchanged ---
     # Dynamically adjust size based on system configuration
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        width = min(950, int(screen_width * 0.8))
        height = min(800, int(screen_height * 0.8))
        root.geometry(f"{width}x{height}")

        self.root = root
        self.root.title("Text to Excel Converter and Excel Split Tool")
        if platform.system() == "Windows":
            default_font = ("Segoe UI", 10)
        else:
            default_font = ("Arial", 11)
        self.root.option_add("*Font", default_font)

        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        # Outer frame, canvas, and scrollbar for full app scrollability
        self.outer_frame = tk.Frame(root)
        self.outer_frame.grid(row=0, column=0, sticky="nsew")
        self.outer_frame.grid_rowconfigure(0, weight=1)
        self.outer_frame.grid_columnconfigure(0, weight=1)

        self.outer_canvas = Canvas(self.outer_frame, borderwidth=0)
        self.vscrollbar = Scrollbar(self.outer_frame, orient="vertical", command=self.outer_canvas.yview)
        self.hscrollbar = Scrollbar(self.outer_frame, orient="horizontal", command=self.outer_canvas.xview)
        self.outer_canvas.configure(yscrollcommand=self.vscrollbar.set, xscrollcommand=self.hscrollbar.set)
        self.vscrollbar.grid(row=0, column=1, sticky="ns")
        self.hscrollbar.grid(row=1, column=0, sticky="ew")
        self.outer_canvas.grid(row=0, column=0, sticky="nsew")

        self.content_frame = tk.Frame(self.outer_canvas)
        self.content_frame_id = self.outer_canvas.create_window((0, 0), window=self.content_frame, anchor="nw")

        def on_frame_configure(event):
            self.outer_canvas.configure(scrollregion=self.outer_canvas.bbox("all"))
        def on_canvas_configure(event):
            canvas_width = event.width
            self.outer_canvas.itemconfig(self.content_frame_id, width=canvas_width)
        self.content_frame.bind("<Configure>", on_frame_configure)
        self.outer_canvas.bind("<Configure>", on_canvas_configure)

        # Mousewheel scrolling for the whole frame
        def _on_mousewheel(event):
            self.outer_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self.outer_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # --- Stage 1: Text to Single Excel Sheet ---
        self.frame_stage1 = LabelFrame(self.content_frame, text="Stage 1: Text to Single Excel Sheet", padx=20, pady=10)
        self.frame_stage1.pack(pady=10, padx=20, fill="x", expand=True)

        tk.Label(self.frame_stage1, text="Input Text File:").grid(row=0, column=0, sticky="e", pady=5)
        self.input_text_entry = tk.Entry(self.frame_stage1, width=40)
        self.input_text_entry.grid(row=0, column=1, sticky="ew", pady=5)
        self.browse_text_button = tk.Button(self.frame_stage1, text="Browse...", command=self.select_input_text_file)
        self.browse_text_button.grid(row=0, column=2, padx=5, pady=5)

        tk.Label(self.frame_stage1, text="Output Excel File (Single):").grid(row=1, column=0, sticky="e", pady=5)
        self.output_single_excel_entry = tk.Entry(self.frame_stage1, width=40)
        self.output_single_excel_entry.grid(row=1, column=1, sticky="ew", pady=5)
        self.browse_output_single_button = tk.Button(self.frame_stage1, text="Browse...", command=self.select_output_single_excel_file)
        self.browse_output_single_button.grid(row=1, column=2, padx=5, pady=5)

        tk.Label(self.frame_stage1, text="Delimiter:").grid(row=2, column=0, sticky="e", pady=5)
        self.delimiter_entry = tk.Entry(self.frame_stage1, width=10)
        self.delimiter_entry.insert(0, ",")
        self.delimiter_entry.grid(row=2, column=1, sticky="w", pady=5)

        stage1_button_frame = tk.Frame(self.frame_stage1)
        stage1_button_frame.grid(row=3, column=0, columnspan=3, pady=15)
        stage1_button_frame.columnconfigure(0, weight=1)
        stage1_button_frame.columnconfigure(1, weight=1)

        self.convert_single_button = tk.Button(
            stage1_button_frame,
            text="Convert (Skip 1st/Last Row)",
            width=30,
            bg="#0078D7",
            fg="white",
            command=self.run_stage1_conversion_skip_rows
        )
        self.convert_single_button.grid(row=0, column=0, padx=5)

        self.convert_full_button = tk.Button(
            stage1_button_frame,
            text="Convert (Keep All Rows)",
            width=30,
            bg="#0078D0",
            fg="white",
            command=self.run_stage1_conversion_full
        )
        self.convert_full_button.grid(row=0, column=1, padx=5)

        self.frame_stage1.columnconfigure(1, weight=1)

        # --- Stage 2: Split Excel by Column Groups ---
        self.frame_stage2 = LabelFrame(self.content_frame, text="Stage 2: Split Excel by Column Groups", padx=20, pady=10)
        self.frame_stage2.pack(pady=10, padx=20, fill="x", expand=True)

        tk.Label(self.frame_stage2, text="Input Excel File:").grid(row=0, column=0, sticky="e", pady=5)
        self.input_split_excel_entry = tk.Entry(self.frame_stage2, width=40)
        self.input_split_excel_entry.grid(row=0, column=1, sticky="ew", pady=5)
        self.browse_split_excel_button = tk.Button(self.frame_stage2, text="Browse...", command=self.select_input_split_excel_file)
        self.browse_split_excel_button.grid(row=0, column=2, padx=5, pady=5)

        tk.Label(self.frame_stage2, text="Output Folder:").grid(row=1, column=0, sticky="e", pady=5)
        self.output_split_folder_entry = tk.Entry(self.frame_stage2, width=40)
        self.output_split_folder_entry.grid(row=1, column=1, sticky="ew", pady=5)
        self.browse_output_split_folder_button = tk.Button(self.frame_stage2, text="Browse...", command=self.select_output_split_folder)
        self.browse_output_split_folder_button.grid(row=1, column=2, padx=5, pady=5)

        tk.Label(self.frame_stage2, text="Defined Column Groups:").grid(row=2, column=0, sticky="nw", pady=5)
        self.split_groups_listbox_frame = tk.Frame(self.frame_stage2)
        self.split_groups_listbox_frame.grid(row=2, column=1, sticky="nsew", pady=5)
        self.split_groups_listbox = Listbox(self.split_groups_listbox_frame, height=5, width=50)
        self.split_groups_listbox.pack(side="left", fill="both", expand=True)
        self.split_groups_scrollbar = Scrollbar(self.split_groups_listbox_frame, command=self.split_groups_listbox.yview)
        self.split_groups_scrollbar.pack(side="right", fill="y")
        self.split_groups_listbox.config(yscrollcommand=self.split_groups_scrollbar.set)
        self.split_groups_listbox.bind("<<ListboxSelect>>", self.on_group_select)

        split_group_button_frame = tk.Frame(self.frame_stage2)
        split_group_button_frame.grid(row=3, column=0, columnspan=3, pady=5)
        self.add_group_button = tk.Button(split_group_button_frame, text="Add Group", state=tk.DISABLED, command=self.add_column_group)
        self.add_group_button.grid(row=0, column=0, padx=5)
        self.edit_group_button = tk.Button(split_group_button_frame, text="Edit Selected Group", state=tk.DISABLED, command=self.edit_selected_group)
        self.edit_group_button.grid(row=0, column=1, padx=5)
        self.remove_group_button = tk.Button(split_group_button_frame, text="Remove Selected Group", state=tk.DISABLED, command=self.remove_selected_group)
        self.remove_group_button.grid(row=0, column=2, padx=5)

        self.group_definition_frame = LabelFrame(self.frame_stage2, text="Define/Edit Column Group", padx=10, pady=10)

        tk.Label(self.group_definition_frame, text="Output File Name (without .xlsx):").grid(row=0, column=0, sticky="e", padx=5, pady=5)
        self.output_file_name_entry = tk.Entry(self.group_definition_frame, width=40)
        self.output_file_name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew", columnspan=2)

        tk.Label(self.group_definition_frame, text="Select Columns for this Group:").grid(row=1, column=0, sticky="nw", padx=5, pady=5)
        self.headers_checkbox_container = tk.Frame(self.group_definition_frame)
        self.headers_checkbox_container.grid(row=1, column=1, padx=5, pady=5, sticky="nsew", columnspan=2)

        self.headers_canvas = Canvas(self.headers_checkbox_container)
        self.headers_scrollbar = Scrollbar(self.headers_checkbox_container, orient="vertical", command=self.headers_canvas.yview)
        self.headers_checkbox_frame = tk.Frame(self.headers_canvas)
        self.headers_canvas.create_window((0, 0), window=self.headers_checkbox_frame, anchor="nw")
        self.headers_canvas.configure(yscrollcommand=self.headers_scrollbar.set)
        self.headers_scrollbar.pack(side="right", fill="y")
        self.headers_canvas.pack(side="left", fill="both", expand=True)
        def on_headers_frame_configure(event):
            self.headers_canvas.configure(scrollregion=self.headers_checkbox_frame.bbox("all"))
        self.headers_checkbox_frame.bind("<Configure>", on_headers_frame_configure)

        select_buttons_frame = tk.Frame(self.group_definition_frame)
        select_buttons_frame.grid(row=2, column=1, columnspan=2, pady=5)
        self.select_all_button = tk.Button(select_buttons_frame, text="Select All", command=self.select_all_headers_checkboxes)
        self.select_all_button.grid(row=0, column=0, padx=5)
        self.deselect_all_button = tk.Button(select_buttons_frame, text="Deselect All", command=self.deselect_all_headers_checkboxes)
        self.deselect_all_button.grid(row=0, column=1, padx=5)

        inline_button_frame = tk.Frame(self.group_definition_frame)
        inline_button_frame.grid(row=3, column=0, columnspan=3, pady=10)
        self.save_group_button = tk.Button(inline_button_frame, text="Save Group", command=self.save_column_group)
        self.save_group_button.grid(row=0, column=0, padx=5)
        self.cancel_group_button = tk.Button(inline_button_frame, text="Cancel", command=self.cancel_column_group_edit)
        self.cancel_group_button.grid(row=0, column=1, padx=5)

        self.group_definition_frame.columnconfigure(1, weight=1)
        self.group_definition_frame.rowconfigure(1, weight=1)

        self.perform_split_button = tk.Button(
            self.frame_stage2,
            text="Perform Split",
            width=50,
            bg="#28A745",
            fg="white",
            state=tk.DISABLED,
            command=self.perform_column_group_split
        )
        self.perform_split_button.grid(row=4, column=0, columnspan=3, pady=15)

        self.dataiq_button_stage2 = tk.Button(
            self.frame_stage2,
            text="DataIQ",
            width=20,
            bg="#4285F4",
            fg="white",
            command=self.open_dataiq_url
        )
        self.dataiq_button_stage2.grid(row=5, column=0, columnspan=3, pady=(0, 10))

        self.frame_stage2.columnconfigure(1, weight=1)
        self.frame_stage2.rowconfigure(2, weight=1)

        self.defined_column_groups = []
        self.all_loaded_headers = []
        self.header_checkbox_vars = []
        self.editing_group_index = None

        # --- Stage 3: Search Value in Excel Column ---
        self.frame_stage3 = LabelFrame(self.content_frame, text="Stage 3: Search Value in Excel Column", padx=20, pady=10)
        self.frame_stage3.pack(pady=10, padx=20, fill="x", expand=True)

        tk.Label(self.frame_stage3, text="Input Excel File (for Search):").grid(row=0, column=0, sticky="e", pady=5)
        self.input_search_excel_entry = tk.Entry(self.frame_stage3, width=40)
        self.input_search_excel_entry.grid(row=0, column=1, sticky="ew", pady=5)
        self.browse_search_excel_button = tk.Button(self.frame_stage3, text="Browse...", command=lambda: self.input_search_excel_entry.insert(0, filedialog.askopenfilename(title="Select Input Excel File")))
        self.browse_search_excel_button.grid(row=0, column=2, padx=5, pady=5)

        self.load_search_headers_button = tk.Button(self.frame_stage3, text="Load Columns", command=self.load_search_excel_columns)
        self.load_search_headers_button.grid(row=1, column=0, columnspan=3, pady=5)

        tk.Label(self.frame_stage3, text="Select Column:").grid(row=2, column=0, sticky="e", pady=5)
        self.search_column_combobox = ttk.Combobox(self.frame_stage3, width=37, state="disabled")
        self.search_column_combobox.grid(row=2, column=1, sticky="ew", pady=5)

        tk.Label(self.frame_stage3, text="Value to Search:").grid(row=3, column=0, sticky="e", pady=5)
        self.search_value_entry = tk.Entry(self.frame_stage3, width=40, state="disabled")
        self.search_value_entry.grid(row=3, column=1, sticky="ew", pady=5)

        self.search_button = tk.Button(
            self.frame_stage3,
            text="Search",
            width=20,
            bg="#17A2B8",
            fg="white",
            state="disabled",
            command=self.perform_search
        )
        self.search_button.grid(row=3, column=2, padx=5, pady=5)

        tk.Label(self.frame_stage3, text="Search Results:").grid(row=4, column=0, sticky="nw", pady=5)
        self.search_results_frame = tk.Frame(self.frame_stage3)
        self.search_results_frame.grid(row=4, column=1, columnspan=2, sticky="nsew", pady=5)
        self.search_results_text = tk.Text(self.search_results_frame, height=10, width=60, state="disabled", wrap="none")
        self.search_results_text.pack(side="left", fill="both", expand=True)
        self.search_results_scrollbar = Scrollbar(self.search_results_frame, command=self.search_results_text.yview)
        self.search_results_scrollbar.pack(side="right", fill="y")
        self.search_results_text.config(yscrollcommand=self.search_results_scrollbar.set)

        self.frame_stage3.columnconfigure(1, weight=1)
        self.frame_stage3.rowconfigure(4, weight=1)

    # --- Stage 1 methods ---
    def select_input_text_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Input Text File (Stage 1)",
            filetypes=[("Text Files", "*.txt *.csv"), ("All Files", "*.*")]
        )
        if file_path:
            self.input_text_entry.delete(0, tk.END)
            self.input_text_entry.insert(0, file_path)

    def select_output_single_excel_file(self):
        file_path = filedialog.asksaveasfilename(
            title="Save Output Single Sheet Excel As (Stage 1)",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if file_path:
            self.output_single_excel_entry.delete(0, tk.END)
            self.output_single_excel_entry.insert(0, file_path)

    def run_stage1_conversion_skip_rows(self):
        input_file = self.input_text_entry.get()
        output_file = self.output_single_excel_entry.get()
        delimiter = self.delimiter_entry.get()
        if not input_file:
            messagebox.showerror("Input Error", "Please select an Input Text File (Stage 1).")
            return
        if not output_file:
            messagebox.showerror("Output Error", "Please specify an Output Single Sheet Excel File (Stage 1).")
            return
        if not delimiter:
            messagebox.showerror("Input Error", "Please provide a Delimiter.")
            return
        self.convert_single_button.config(state=tk.DISABLED)
        self.convert_full_button.config(state=tk.DISABLED)
        self.root.update_idletasks()
        success, msg = self.convert_text_to_excel_skip_first_last(input_file, output_file, delimiter)
        self.convert_single_button.config(state=tk.NORMAL)
        self.convert_full_button.config(state=tk.NORMAL)
        if success:
            messagebox.showinfo("Stage 1 Success", msg)
        else:
            messagebox.showerror("Stage 1 Failed", msg)

    def run_stage1_conversion_full(self):
        input_file = self.input_text_entry.get()
        output_file = self.output_single_excel_entry.get()
        delimiter = self.delimiter_entry.get()
        if not input_file:
            messagebox.showerror("Input Error", "Please select an Input Text File (Stage 1).")
            return
        if not output_file:
            messagebox.showerror("Output Error", "Please specify an Output Single Sheet Excel File (Stage 1).")
            return
        if not delimiter:
            messagebox.showerror("Input Error", "Please provide a Delimiter.")
            return
        self.convert_single_button.config(state=tk.DISABLED)
        self.convert_full_button.config(state=tk.DISABLED)
        self.root.update_idletasks()
        success, msg = self.convert_text_to_excel_full(input_file, output_file, delimiter)
        self.convert_single_button.config(state=tk.NORMAL)
        self.convert_full_button.config(state=tk.NORMAL)
        if success:
            messagebox.showinfo("Stage 1 Success", msg)
        else:
            messagebox.showerror("Stage 1 Failed", msg)

    def convert_text_to_excel_skip_first_last(self, input_file, output_file, delimiter):
        try:
            with open(input_file, "r", encoding="utf-8") as f:
                lines = f.readlines()
            if len(lines) <= 2:
                return False, "Not enough lines in the text file."
            lines = lines[1:-1]
            rows = [line.strip().split(delimiter) for line in lines]
            rows = [[cell.strip('"') for cell in row] for row in rows]
            df = pd.DataFrame(rows)
            df.to_excel(output_file, index=False, header=False)
            return True, f"File converted and saved to {output_file}"
        except Exception as e:
            return False, str(e)

    def convert_text_to_excel_full(self, input_file, output_file, delimiter):
        try:
            with open(input_file, "r", encoding="utf-8") as f:
                lines = f.readlines()
            rows = [line.strip().split(delimiter) for line in lines]
            rows = [[cell.strip('"') for cell in row] for row in rows]
            df = pd.DataFrame(rows)
            df.to_excel(output_file, index=False, header=False)
            return True, f"File converted and saved to {output_file}"
        except Exception as e:
            return False, str(e)

    # --- Stage 2 methods (split, group management, scrollbars, etc.) ---
    def select_input_split_excel_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Input Excel File (Stage 2)",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if file_path:
            self.input_split_excel_entry.delete(0, tk.END)
            self.input_split_excel_entry.insert(0, file_path)
            self.load_split_excel_headers()
            self.clear_defined_groups()
            self.hide_group_definition_frame()

    def select_output_split_folder(self):
        folder_path = filedialog.askdirectory(
            title="Select Output Folder for Split Files (Stage 2)"
        )
        if folder_path:
            self.output_split_folder_entry.delete(0, tk.END)
            self.output_split_folder_entry.insert(0, folder_path)

    def load_split_excel_headers(self):
        input_excel_file = self.input_split_excel_entry.get()
        self.all_loaded_headers = []
        self.add_group_button.config(state=tk.DISABLED)
        self.edit_group_button.config(state=tk.DISABLED)
        self.remove_group_button.config(state=tk.DISABLED)
        self.perform_split_button.config(state=tk.DISABLED)
        self.clear_defined_groups()
        self.clear_header_checkboxes()
        self.hide_group_definition_frame()
        if not input_excel_file:
            return
        if not os.path.exists(input_excel_file):
            messagebox.showwarning("File Not Found", f"Input Excel file not found: {input_excel_file}")
            return
        try:
            headers_df = pd.read_excel(input_excel_file, sheet_name=0, nrows=0)
            headers = headers_df.columns.tolist()
            if headers:
                self.all_loaded_headers = headers
                self.create_header_checkboxes(headers)
                self.add_group_button.config(state=tk.NORMAL)
            else:
                messagebox.showwarning("No Headers Found", f"Could not detect headers in the first sheet of Excel file: {input_excel_file}.\nCheck if the first row contains headers.")
        except Exception as e:
            messagebox.showerror("Error Loading Excel Headers", str(e))

    def create_header_checkboxes(self, headers):
        self.clear_header_checkboxes()
        self.header_checkbox_vars = []
        for i, header in enumerate(headers):
            var = BooleanVar()
            cb = Checkbutton(self.headers_checkbox_frame, text=header, variable=var, anchor="w")
            cb.grid(row=i, column=0, sticky="w")
            self.header_checkbox_vars.append(var)
        self.headers_checkbox_frame.update_idletasks()
        self.on_headers_frame_configure(None)

    def clear_header_checkboxes(self):
        for widget in self.headers_checkbox_frame.winfo_children():
            widget.destroy()
        self.header_checkbox_vars = []

    def select_all_headers_checkboxes(self):
        for var in self.header_checkbox_vars:
            var.set(True)

    def deselect_all_headers_checkboxes(self):
        for var in self.header_checkbox_vars:
            var.set(False)

    def clear_defined_groups(self):
        self.defined_column_groups = []
        self.update_groups_listbox()
        self.edit_group_button.config(state=tk.DISABLED)
        self.remove_group_button.config(state=tk.DISABLED)
        self.perform_split_button.config(state=tk.DISABLED)

    def update_groups_listbox(self):
        self.split_groups_listbox.delete(0, tk.END)
        if not self.defined_column_groups:
            self.split_groups_listbox.insert(tk.END, "No column groups defined yet.")
            self.edit_group_button.config(state=tk.DISABLED)
            self.remove_group_button.config(state=tk.DISABLED)
            self.perform_split_button.config(state=tk.DISABLED)
            return
        for i, (output_file_name, columns) in enumerate(self.defined_column_groups):
            display_text = f"Group {i+1}: {', '.join(columns)} -> {output_file_name}.xlsx"
            self.split_groups_listbox.insert(tk.END, display_text)
        self.perform_split_button.config(state=tk.NORMAL)
        self.split_groups_listbox.config(state=tk.NORMAL)

    def show_group_definition_frame(self):
        self.group_definition_frame.grid(row=6, column=0, columnspan=3, sticky="ew", pady=10)
        self.add_group_button.config(state=tk.DISABLED)
        self.edit_group_button.config(state=tk.DISABLED)
        self.remove_group_button.config(state=tk.DISABLED)
        self.perform_split_button.config(state=tk.DISABLED)
        self.split_groups_listbox.config(state=tk.DISABLED)
        self.output_split_folder_entry.config(state=tk.DISABLED)
        self.browse_output_split_folder_button.config(state=tk.DISABLED)

    def hide_group_definition_frame(self):
        self.group_definition_frame.grid_forget()
        self.split_groups_listbox.config(state=tk.NORMAL)
        self.add_group_button.config(state=tk.NORMAL)
        self.output_split_folder_entry.config(state=tk.NORMAL)
        self.browse_output_split_folder_button.config(state=tk.NORMAL)
        self.on_group_select(None)

    def add_column_group(self):
        if not self.all_loaded_headers:
            messagebox.showwarning("Headers Not Loaded", "Please load headers from the Excel file first.")
            return
        self.editing_group_index = None
        self.output_file_name_entry.delete(0, tk.END)
        self.deselect_all_headers_checkboxes()
        self.group_definition_frame.config(text="Define New Column Group")
        self.show_group_definition_frame()

    def edit_selected_group(self):
        selected_indices = self.split_groups_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Selection Error", "Please select a column group from the list to edit.")
            return
        self.editing_group_index = selected_indices[0]
        output_file_name, columns = self.defined_column_groups[self.editing_group_index]
        self.output_file_name_entry.delete(0, tk.END)
        self.output_file_name_entry.insert(0, output_file_name)
        self.deselect_all_headers_checkboxes()
        for col in columns:
            try:
                index = self.all_loaded_headers.index(col)
                self.header_checkbox_vars[index].set(True)
            except ValueError:
                pass
        self.group_definition_frame.config(text=f"Edit Column Group {self.editing_group_index + 1}")
        self.show_group_definition_frame()

    def save_column_group(self):
        output_file_name = self.output_file_name_entry.get().strip()
        selected_columns = [self.all_loaded_headers[i] for i, var in enumerate(self.header_checkbox_vars) if var.get()]
        if not output_file_name:
            messagebox.showwarning("Input Error", "Please specify an output file name.")
            return
        output_file_name = re.sub(r'[^\w\s.-]', '', output_file_name)
        output_file_name = output_file_name.replace(' ', '_')
        if not selected_columns:
            messagebox.showwarning("Selection Error", "Please select at least one column for this group.")
            return
        if self.editing_group_index is None:
            self.defined_column_groups.append((output_file_name, selected_columns))
        else:
            self.defined_column_groups[self.editing_group_index] = (output_file_name, selected_columns)
        self.split_groups_listbox.config(state=tk.NORMAL)
        self.update_groups_listbox()
        self.hide_group_definition_frame()
        self.editing_group_index = None

    def cancel_column_group_edit(self):
        self.hide_group_definition_frame()
        self.editing_group_index = None

    def remove_selected_group(self):
        selected_indices = self.split_groups_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Selection Error", "Please select a column group from the list to remove.")
            return
        group_index = selected_indices[0]
        del self.defined_column_groups[group_index]
        self.update_groups_listbox()

    def perform_column_group_split(self):
        input_excel_file = self.input_split_excel_entry.get()
        output_folder = self.output_split_folder_entry.get()
        if not input_excel_file:
            messagebox.showerror("Input Error", "Please select an Input Excel File (Stage 2).")
            return
        if not output_folder:
            messagebox.showerror("Input Error", "Please specify an Output Folder (Stage 2).")
            return
        if not self.defined_column_groups:
            messagebox.showwarning("No Groups Defined", "Please define at least one column group to perform the split.")
            return
        if not os.path.exists(output_folder):
            try:
                os.makedirs(output_folder)
            except Exception as e:
                messagebox.showerror("Folder Creation Error", f"Could not create output folder: {e}")
                return
        output_file_names = [group[0] for group in self.defined_column_groups]
        column_groups_list = [group[1] for group in self.defined_column_groups]
        self.add_group_button.config(state=tk.DISABLED)
        self.edit_group_button.config(state=tk.DISABLED)
        self.remove_group_button.config(state=tk.DISABLED)
        self.perform_split_button.config(state=tk.DISABLED)
        self.split_groups_listbox.config(state=tk.DISABLED)
        self.output_split_folder_entry.config(state=tk.DISABLED)
        self.browse_output_split_folder_button.config(state=tk.DISABLED)
        self.convert_single_button.config(state=tk.DISABLED)
        self.convert_full_button.config(state=tk.DISABLED)
        self.dataiq_button_stage2.config(state=tk.DISABLED)
        self.root.update_idletasks()
        try:
            df = pd.read_excel(input_excel_file, sheet_name=0, header=0, dtype=str)
            split_count = 0
            for output_file_name, columns_to_include in zip(output_file_names, column_groups_list):
                output_file_path = os.path.join(output_folder, f"{output_file_name}.xlsx")
                try:
                    missing_cols = [col for col in columns_to_include if col not in df.columns]
                    if missing_cols:
                        messagebox.showwarning("Missing Columns", f"Skipping group for '{output_file_name}.xlsx' due to missing columns in the first sheet: {', '.join(missing_cols)}")
                        continue
                    df_subset = df[columns_to_include]
                    df_subset.to_excel(output_file_path, index=False)
                    if os.path.exists(output_file_path) and os.path.getsize(output_file_path) > 100:
                        wb = load_workbook(output_file_path)
                        ws = wb.active
                        text_fmt = '@'
                        for row in ws.iter_rows():
                            for cell in row:
                                cell.number_format = text_fmt
                        wb.save(output_file_path)
                    split_count += 1
                except Exception as save_error:
                    messagebox.showwarning("Save Error", f"Could not save group to '{output_file_path}': {save_error}")
            if split_count > 0:
                messagebox.showinfo("Split Success", f"Successfully split Excel file into {split_count} files in folder: {output_folder}")
            else:
                messagebox.showwarning("Split Completed", f"Split operation completed, but no files were successfully created in folder: {output_folder}")
        except FileNotFoundError:
            messagebox.showerror("File Not Found", f"Input Excel file not found at {input_excel_file}")
        except Exception as e:
            messagebox.showerror("Split Failed", str(e))
        finally:
            self.convert_single_button.config(state=tk.NORMAL)
            self.convert_full_button.config(state=tk.NORMAL)
            self.dataiq_button_stage2.config(state=tk.NORMAL)
            self.input_split_excel_entry.config(state=tk.NORMAL)
            self.browse_split_excel_button.config(state=tk.NORMAL)
            self.output_split_folder_entry.config(state=tk.NORMAL)
            self.browse_output_split_folder_button.config(state=tk.NORMAL)
            if self.all_loaded_headers:
                self.add_group_button.config(state=tk.NORMAL)
            if self.defined_column_groups:
                self.perform_split_button.config(state=tk.NORMAL)
            self.split_groups_listbox.config(state=tk.NORMAL)

    # --- Stage 3 methods ---
    def load_search_excel_columns(self):
        input_excel_file = self.input_search_excel_entry.get()
        self.search_column_combobox.set('')
        self.search_column_combobox['values'] = []
        self.search_column_combobox.config(state="disabled")
        self.search_value_entry.delete(0, tk.END)
        self.search_value_entry.config(state="disabled")
        self.search_button.config(state="disabled")
        self.search_results_text.config(state="normal")
        self.search_results_text.delete(1.0, tk.END)
        self.search_results_text.insert(tk.END, "Load an Excel file to see columns.")
        self.search_results_text.config(state="disabled")
        if not input_excel_file:
            messagebox.showwarning("Input Error", "Please select an Input Excel File for Search (Stage 3).")
            return
        if not os.path.exists(input_excel_file):
            messagebox.showwarning("File Not Found", f"Input Excel file not found: {input_excel_file}")
            return
        try:
            headers_df = pd.read_excel(input_excel_file, nrows=0)
            headers = headers_df.columns.tolist()
            if headers:
                self.search_column_combobox['values'] = headers
                self.search_column_combobox.config(state="readonly")
                self.search_results_text.config(state="normal")
                self.search_results_text.delete(1.0, tk.END)
                self.search_results_text.insert(tk.END, f"Columns loaded. Select a column and enter a value to search.")
                self.search_results_text.config(state="disabled")
                self.search_value_entry.config(state="normal")
                self.search_button.config(state="normal")
            else:
                messagebox.showwarning("No Headers Found", f"Could not detect headers in Excel file: {input_excel_file}.\nCheck if the first row contains headers.")
                self.search_results_text.config(state="normal")
                self.search_results_text.delete(1.0, tk.END)
                self.search_results_text.insert(tk.END, "No headers found in the selected file.")
                self.search_results_text.config(state="disabled")
        except Exception as e:
            messagebox.showerror("Error Loading Excel Headers", str(e))
            self.search_results_text.config(state="normal")
            self.search_results_text.delete(1.0, tk.END)
            self.search_results_text.insert(tk.END, f"An error occurred during the search: {e}")
            self.search_results_text.config(state="disabled")

    def perform_search(self):
        input_excel_file = self.input_search_excel_entry.get()
        selected_column = self.search_column_combobox.get()
        search_value = self.search_value_entry.get()
        self.search_results_text.config(state="normal")
        self.search_results_text.delete(1.0, tk.END)
        self.search_results_text.config(state="disabled")
        if not input_excel_file:
            messagebox.showwarning("Input Error", "Please select an Input Excel File for Search (Stage 3).")
            return
        if not selected_column:
            messagebox.showwarning("Selection Error", "Please select a column to search in.")
            self.search_results_text.config(state="normal")
            self.search_results_text.insert(tk.END, "Please select a column.")
            self.search_results_text.config(state="disabled")
            return
        if not search_value:
            messagebox.showwarning("Input Error", "Please enter a value to search for.")
            self.search_results_text.config(state="normal")
            self.search_results_text.insert(tk.END, "Please enter a value to search.")
            self.search_results_text.config(state="disabled")
            return
        if not os.path.exists(input_excel_file):
            messagebox.showwarning("File Not Found", f"Input Excel file not found: {input_excel_file}")
            self.search_results_text.config(state="normal")
            self.search_results_text.insert(tk.END, "Error: File not found.")
            self.search_results_text.config(state="disabled")
            return
        try:
            df = pd.read_excel(input_excel_file, dtype=str)
            if selected_column not in df.columns:
                messagebox.showerror("Column Error", f"Selected column '{selected_column}' not found in the Excel file.")
                self.search_results_text.config(state="normal")
                self.search_results_text.insert(tk.END, f"Error: Column '{selected_column}' not found.")
                self.search_results_text.config(state="disabled")
                return
            matching_rows_df = df[df[selected_column].astype(str).str.contains(search_value, case=False, na=False, regex=False)]
            self.search_results_text.config(state="normal")
            self.search_results_text.delete(1.0, tk.END)
            if not matching_rows_df.empty:
                results_string = matching_rows_df.to_string(index=False)
                self.search_results_text.insert(tk.END, results_string)
            else:
                self.search_results_text.insert(tk.END, f"No results found for '{search_value}' in column '{selected_column}'.")
            self.search_results_text.config(state="disabled")
        except Exception as e:
            messagebox.showerror("Search Error", str(e))
            self.search_results_text.config(state="normal")
            self.search_results_text.insert(tk.END, f"An error occurred during the search: {e}")
            self.search_results_text.config(state="disabled")

    def on_group_select(self, event):
        if self.group_definition_frame.winfo_ismapped():
            self.edit_group_button.config(state=tk.DISABLED)
            self.remove_group_button.config(state=tk.DISABLED)
        else:
            if self.split_groups_listbox.curselection():
                self.edit_group_button.config(state=tk.NORMAL)
                self.remove_group_button.config(state=tk.NORMAL)
            else:
                self.edit_group_button.config(state=tk.DISABLED)
                self.remove_group_button.config(state=tk.DISABLED)

    def on_headers_frame_configure(self, event):
        self.headers_canvas.configure(scrollregion=self.headers_checkbox_frame.bbox("all"))

    def open_dataiq_url(self):
        dataiq_url = "https://www.example.com/dataiq"
        try:
            webbrowser.open(dataiq_url)
        except Exception as e:
            messagebox.showerror("Error Opening URL", f"Could not open the DataIQ URL: {e}")

# --- Help Tab ---
class HelpTab(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        tk.Label(self, text="Help Document", font=("Arial", 14, "bold")).pack(pady=8)
        tk.Label(self, text="This application provides two main tools:\n"
                            "- Side-by-Side File Comparison Tool\n"
                            "- Text to Excel Converter and Split Tool\n\n"
                            "For full documentation and examples, click the link below.",
                 justify="left", font=("Arial", 11)).pack(padx=12, pady=4)
        link = tk.Label(self, text=HELP_URL, fg="blue", cursor="hand2", font=("Arial", 10, "underline"))
        link.pack(pady=8)
        link.bind("<Button-1>", lambda e: webbrowser.open_new(HELP_URL))
        tk.Label(self, text="You may also find detailed instructions in the HELP.md file in your repository.",
                 justify="left", font=("Arial", 10)).pack(padx=12, pady=4)

def main():
    root = tk.Tk()
    root.title("All-in-One Excel Tool")
    root.geometry("1200x800")
    notebook = ttk.Notebook(root)
    notebook.pack(fill="both", expand=True)
    sbs_tab = MappingSearchSBSApp(notebook)
    notebook.add(sbs_tab, text="File Comparison Tool")
    excel_tab = ExcelToolApp(notebook)
    notebook.add(excel_tab, text="Text/Excel Split Tool")
    help_tab = HelpTab(notebook)
    notebook.add(help_tab, text="Help")

    menubar = tk.Menu(root)
    helpmenu = tk.Menu(menubar, tearoff=0)
    helpmenu.add_command(label="Open Online Help", command=lambda: webbrowser.open_new(HELP_URL))
    menubar.add_cascade(label="Help", menu=helpmenu)
    root.config(menu=menubar)

    root.mainloop()

if __name__ == "__main__":
    main()
